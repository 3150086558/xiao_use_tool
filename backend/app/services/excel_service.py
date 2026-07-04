import io
import re
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


def build_import_template() -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "导入数据"
    headers = ["日期", "类型", "项目", "子分类", "金额", "账户", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    required = ["必填", "必填", "必填", "选填", "必填", "选填", "选填"]
    for col, text in enumerate(required, 1):
        ws1.cell(row=2, column=col, value=text)

    examples = [
        ["2024-01-15", "支出", "餐饮", "午餐", 35.5, "微信", "公司楼下"],
        ["2024-01-16", "收入", "工资", "", 10000, "银行卡", "1月工资"],
    ]
    for i, row in enumerate(examples, 3):
        for j, val in enumerate(row, 1):
            ws1.cell(row=i, column=j, value=val)

    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[chr(64 + col)].width = 15

    ws2 = wb.create_sheet("字段说明")
    descriptions = [
        ["字段", "说明", "示例"],
        ["日期", "支持 YYYY-MM-DD、YYYY/MM/DD、YYYY年MM月DD日 格式", "2024-01-15"],
        ["类型", "收入 或 支出", "支出"],
        ["项目", "消费或收入的大类", "餐饮"],
        ["子分类", "更细的分类（可空）", "午餐"],
        ["金额", "正数；负数会自动转成支出", "35.5"],
        ["账户", "支付/收入账户（可空）", "微信"],
        ["备注", "其他说明（可空）", "公司楼下"],
    ]
    for i, row in enumerate(descriptions, 1):
        for j, val in enumerate(row, 1):
            cell = ws2.cell(row=i, column=j, value=val)
            if i == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parse_date(s: str) -> date | None:
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    try:
        if isinstance(s, datetime):
            return s.date()
        if isinstance(s, date):
            return s
    except Exception:
        pass
    return None


def parse_import_excel(file_content: bytes) -> tuple[list[dict], list[str]]:
    wb = load_workbook(io.BytesIO(file_content), data_only=True)
    ws = wb.active

    records = []
    errors = []

    type_map = {
        "收入": "income", "income": "income",
        "支出": "expense", "expense": "expense",
    }

    for row_idx in range(3, ws.max_row + 1):
        row = [ws.cell(row=row_idx, column=c).value for c in range(1, 8)]
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        date_val = parse_date(row[0])
        if not date_val:
            errors.append(f"第 {row_idx} 行：日期格式不正确 ({row[0]})")
            continue

        type_raw = str(row[1]).strip() if row[1] else ""
        type_val = type_map.get(type_raw)
        if not type_val:
            errors.append(f"第 {row_idx} 行：类型不正确 ({type_raw})，应为 收入 或 支出")
            continue

        category = str(row[2]).strip() if row[2] else ""
        if not category:
            errors.append(f"第 {row_idx} 行：项目不能为空")
            continue

        sub_category = str(row[3]).strip() if row[3] else ""

        try:
            amount = Decimal(str(row[4])) if row[4] else Decimal("0")
            if amount < 0:
                amount = abs(amount)
                type_val = "expense" if type_val == "income" else "income"
        except (InvalidOperation, ValueError):
            errors.append(f"第 {row_idx} 行：金额格式不正确 ({row[4]})")
            continue

        if amount <= 0:
            errors.append(f"第 {row_idx} 行：金额必须大于 0")
            continue

        account = str(row[5]).strip() if row[5] else ""
        note = str(row[6]).strip() if row[6] else ""

        records.append({
            "record_date": date_val,
            "type": type_val,
            "category": category,
            "sub_category": sub_category,
            "amount": amount,
            "account": account,
            "note": note,
        })

    return records, errors


def export_to_excel(records: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "记账数据"

    headers = ["日期", "类型", "项目", "子分类", "金额", "账户", "备注"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for i, rec in enumerate(records, 2):
        ws.cell(row=i, column=1, value=str(rec.get("record_date", "")))
        ws.cell(row=i, column=2, value="收入" if rec.get("type") == "income" else "支出")
        ws.cell(row=i, column=3, value=rec.get("category", ""))
        ws.cell(row=i, column=4, value=rec.get("sub_category", ""))
        ws.cell(row=i, column=5, value=float(rec.get("amount", 0)))
        ws.cell(row=i, column=6, value=rec.get("account", ""))
        ws.cell(row=i, column=7, value=rec.get("note", ""))

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_to_csv(records: list[dict]) -> str:
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["日期", "类型", "项目", "子分类", "金额", "账户", "备注"])
    for rec in records:
        writer.writerow([
            str(rec.get("record_date", "")),
            "收入" if rec.get("type") == "income" else "支出",
            rec.get("category", ""),
            rec.get("sub_category", ""),
            float(rec.get("amount", 0)),
            rec.get("account", ""),
            rec.get("note", ""),
        ])
    return output.getvalue()
